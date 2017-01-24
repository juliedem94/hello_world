import json
from openpyxl import load_workbook

def loadXLSX(name, list_name):
    wb = load_workbook(name + '.xlsx', data_only = True)
    ws = wb[list_name]
    col = 1
    row = 1    
    fields = []
    field = True
    while field:
        field = ws.cell(row = row, column = col).value    
        if field:
            fields.append(field)
            col += 1
    
    data1 = []
    row += 1
    not_empty = True
    while not_empty:
        entry = {}
        not_empty = False
        for i, field in enumerate(fields):
            col = i + 1
            val = ws.cell(row = row, column = col).value            
            entry[field] = val
            if val:
                not_empty = True
        if not_empty:
            data1.append(entry)
            row += 1        
    
    return data1
    
def jsonSave(file_name, data):
    file = open(file_name, 'w', encoding = 'utf8')    
    json.dump(data, file, ensure_ascii = False, indent = 2)
    file.close()
    
def jsonLoad(file_name):
    file = open(file_name, 'r', encoding = 'utf8')
    return json.load(file)
    