# -*- coding: utf-8 -*-
"""
Created on Sun Jan 22 16:15:03 2017

@author: РС
"""

import vk
import time
import pickle

from openpyxl import load_workbook
from openpyxl import Workbook



token = 'a3111388be7e62da597da4f248d5e4eb59880bc749eb487ae163892e077ef5af41dd12144a204c09849f7'
session = vk.Session(access_token = token) 
api = vk.API(session, v = '5.45')


def loadXLSX(name, list_name):
    wb = load_workbook(name + '.xlsx', data_only = True)
    ws = wb[list_name]
    col = 1
    row = 1    
    fields = []
    field = True
    while field:
        field = ws.cell(row = row, column = col).value    
        if field:
            fields.append(field)
            col += 1
    
    data1 = []
    row += 1
    not_empty = True
    while not_empty:
        entry = {}
        not_empty = False
        for i, field in enumerate(fields):
            col = i + 1
            val = ws.cell(row = row, column = col).value            
            entry[field] = val
            if val:
                not_empty = True
        if not_empty:
            data1.append(entry)
            row += 1        
    
    return data1
    

school_index = loadXLSX('C:/Users/РС/Desktop/school', 'Лист1')


h=0
curr_school = 0


while h<34:
    
    print('----------------------School '+str(school_index[h]['School'])+'----------------------------------')
    
    curr_school=school_index[h]['School']

    vk_data = api.users.search(school = curr_school,count=1000, birth_year = 1996)
    users = vk_data['items'] 
    time.sleep(0.35)
    
    
    users_id = [s['id'] for s in users]
    #print(users_id)
    
    pickle.dump(users_id, open('C:/Users/РС/Desktop/код/hello_world/data/' + str(curr_school) +'.p', "wb" ) )
        
 
    h=h+1 
    












