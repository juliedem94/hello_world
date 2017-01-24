# -*- coding: utf-8 -*-
"""
Created on Sun Jan 22 16:15:03 2017

@author: РС
"""

import vk
import time
import pickle
import math

from openpyxl import load_workbook


token = 'dc8f832ac59fe7e40088dec5ed0a3eb3ed314797f27b22d89f8bc123d4f660afdb03f86be24a0afa87635'
session = vk.Session(access_token = token) 
api = vk.API(session, v = '5.45')

def get_friends(users):
    data = []
    max_num = math.floor(len(users) / 20) + 1
    for i in range(0, max_num):
        new_users = users[i * 20:i * 20 + 20]
        sub_data = []
        code_parts = []
        for user in new_users:
            code_parts.append('API.friends.get({"user_id": ' + str(user)+'}).items')
        code_string = 'return [' + ','.join(code_parts) +'];'  
        print(code_string)
        done = False
        error_count = 0
        while not done:
            try:
                time.sleep(0.335)   
                sub_data = api.execute(code = code_string, timeout = 60)
                done = True
            except:                
                error_count += 1
                print("Error!")
                time.sleep(1)
                if error_count > 5:
                    print('Too many errors in a row :(')
                    done = True
            
        data = data + sub_data
        print(i + 1, 'out of', max_num)
            
    return data

def loadXLSX(name, list_name):
    wb = load_workbook(name + '.xlsx', data_only = True)
    ws = wb[list_name]
    col = 1
    row = 1    
    fields = []
    field = True
    while field:
        field = ws.cell(row = row, column = col).value    
        if field:
            fields.append(field)
            col += 1
    
    data1 = []
    row += 1
    not_empty = True
    while not_empty:
        entry = {}
        not_empty = False
        for i, field in enumerate(fields):
            col = i + 1
            val = ws.cell(row = row, column = col).value            
            entry[field] = val
            if val:
                not_empty = True
        if not_empty:
            data1.append(entry)
            row += 1        
    
    return data1
    
school_index = loadXLSX('C:/Users/РС/Desktop/school', 'Лист1')    
    
h=0
curr_school = 0



while h<34:
    
    print('----------------------School '+str(school_index[h]['School'])+'----------------------------------')
    
    curr_school=school_index[h]['School']
    
    users_id=pickle.load(open('C:/Users/РС/Desktop/код/hello_world/data/' + str(curr_school) + '.p', "rb" ))
    
    users_friends_sum =[] 
   
    users_friends_sum = get_friends(users_id)
    
    
    pickle.dump(users_friends_sum, open( 'C:/Users/РС/Desktop/код/hello_world/data/friends_school_' + str(curr_school) + '.p', "wb" )) 
                                             
    h=h+1
    print('------------------done-------------------')




