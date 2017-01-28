# -*- coding: utf-8 -*-
"""
Created on Wed Jan 11 23:22:17 2017

@author: РС
"""

import vk
import time



token = '6c09c77ed5bd767484d66021630c6adfc18ac84ed1edea03218a3f2a9bb6489d3ab570a10dc67d8b0d99a'
session = vk.Session(access_token = token) 
api = vk.API(session, v = '5.45')

#%%
'''
Теперь можно использовать методы API ВКонтакте (https://vk.com/dev/methods)
'''


from openpyxl import load_workbook

#ф-я, которая убрает тех, у кого нет друзей из этой школы
def clean_bots(users):

    data = []
    
    for user in users:
       #print('------------------USER------------------------------------')
       #print(user)
        time.sleep(0.335)  
        user_friends = api.friends.get(user_id=str(user['id']),fields='school')['items']
                                                   

        good_ids = [s['id'] for s in users]
            
          
        for friend in user_friends:
            if friend['id'] in good_ids:
                #список вузов для учеников данной школы п-го года рождения
                time.sleep(0.335)
                good_user_uni = api.users.get(user_ids = str(user['id']), fields = 'universities')
                try:
                    curr_user_uni = good_user_uni[0]['universities']
                    data.append(curr_user_uni)
                
                except:
                    print('у полльзователя не указан университет')

                break
     
       
            
    return data


def loadXLSX(name, list_name):
    wb = load_workbook(name + '.xlsx', data_only = True)
    ws = wb[list_name]
    col = 1
    row = 1    
    fields = []
    field = True
    while field:
        field = ws.cell(row = row, column = col).value    
        if field:
            fields.append(field)
            col += 1
    
    data1 = []
    row += 1
    not_empty = True
    while not_empty:
        entry = {}
        not_empty = False
        for i, field in enumerate(fields):
            col = i + 1
            val = ws.cell(row = row, column = col).value            
            entry[field] = val
            if val:
                not_empty = True
        if not_empty:
            data1.append(entry)
            row += 1        
    
    return data1
    
    
    
    
# Загружаем пользователей из школы № .. до 18 лет
school_index = loadXLSX('C:/Users/РС/Desktop/school', 'Лист1')
ratings = loadXLSX('C:/Users/РС/Desktop/uni', 'Лист1')
h=0
curr_school = 0

uni_rate_total=[]
school_rate_total=[]

while h<34:
    
    print('----------------------School '+str(h)+'----------------------------------')
    
    curr_school=school_index[h]['School']

   
    vk_data = api.users.search(school = curr_school,count=80, birth_year = 1996)
    users = vk_data['items']
    
  
    ###находим баллы ЕГЭ для полученных вузов и считаем среднее
               
    #считаем среднее для универов, полученных ранее
    universities = []
    good_uni = clean_bots(users)
    print(good_uni)
    for j in range (len(good_uni)):
        curr_user_uni_list = good_uni[j]
        for k in range (len(curr_user_uni_list)):
            universities.append(curr_user_uni_list[k]['name'])        
            
    
    universities_overlap = []
    for i in range (len(universities)):
        for a in range (len(ratings)):
            if ratings[a]['University'] ==universities[i]:
                universities_overlap.append(ratings[a])
                
    print(universities_overlap) 
           
    #mean
    
    rate_summ = 0
    for a in range (len(universities_overlap)):
        rate_summ = rate_summ + universities_overlap[a]['Rate']
    

    if rate_summ==0:
        print('Школа ' + str(h) + 'исключена из расчета.')
    else:
        school_rate_total.append(float(school_index[h]['Rate']))
        uni_mean = rate_summ/(len(universities_overlap))
        uni_rate_total.append(uni_mean)
   
  
    h=h+1
    
  
print('----------------------------math---------------------------')    
print(uni_rate_total)   
print(school_rate_total) 
  



      
#корреляция   
from scipy import stats
print ('--------------------------corr------------------------------')
print(stats.pearsonr(uni_rate_total, school_rate_total))

#строим график
import matplotlib.pyplot as plt
plt.scatter(uni_rate_total, school_rate_total)
